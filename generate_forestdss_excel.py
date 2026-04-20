#!/usr/bin/env python3
"""
generate_forestdss_excel.py
Generates forestdss_form.xlsx – ForestDSS Characterization Workbook.

Usage:
    python generate_forestdss_excel.py

Requires: openpyxl  (pip install openpyxl)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName

# ── colour palette ────────────────────────────────────────────────────────
HDR_BG  = "2D5A1B"   # dark green header
HDR_FG  = "FFFFFF"   # white header text
KEY_BG  = "EEEEEE"   # light grey key row
KEY_FG  = "888888"   # grey key text
ALT_BG  = "F4F9F1"   # very light green alternating row
WHT_BG  = "FFFFFF"   # white row

DATA_ROWS = 20        # empty data rows per sheet

# ── controlled vocabularies ───────────────────────────────────────────────
VOCAB: dict[str, list[str]] = {
    "itToolType": [
        "Extension (based on standard software)", "Command line",
        "GUI \u2013 desktop", "Web application", "Mobile device app", "Other",
    ],
    "techReadiness": [
        "Prototype \u2013 no practical application",
        "Prototype \u2013 with practical application",
        "Non-commercial product", "Open-source product",
        "Commercial product", "Other",
    ],
    "targetUser": [
        "Forest enterprise", "Forest contractor",
        "Forest engineering/planning firm", "Forest service/administration",
        "Forest owner/manager", "Conservation manager",
        "Research institution", "Education/training",
        "Stakeholder platform/NGO", "Industry and supply chain actors", "Other",
    ],
    "costs": ["0 \u20ac", "\u2264 100 \u20ac", "\u2264 1,000 \u20ac",
              "\u2264 10,000 \u20ac", "> 10,000 \u20ac"],
    "training": ["\u2264 1 hour", "\u2264 1 day", "\u2264 1 week",
                 "\u2264 1 month", "> 1 month"],
    "realLife": ["0", "\u2264 3", "\u2264 10", "\u2264 30", "\u2264 100", "> 100"],
    "eduUse": [
        "No", "Yes \u2013 presentation/demo only",
        "Yes \u2013 direct use by students",
    ],
    "origin": [
        "Research institution", "PhD project", "MSc project", "BSc project",
        "Public administration", "Private enterprise", "Other",
    ],
    "devTeam": ["1", "2\u20135", "> 5"],
    "devMgmt": ["Scrum", "Kanban", "XP", "None"],
    "yesno": ["Yes", "No", "Unknown"],
    "vegType": ["Empirical", "Process-based", "Hybrid"],
    "spatVeg": ["Tree", "Cohort", "Stand", "Landscape", "> Landscape"],
    "tempVeg": ["Sub-annual", "Annual", "Annual to decadal",
                "> Decadal", "Flexible"],
    "region": [
        "Northern Europe", "Central Europe", "Western Europe",
        "Southern Europe", "Eastern Europe", "Other",
    ],
    "mgmtScale": [
        "Tree", "Stand", "Multi-stands", "Forest management unit",
        "Landscape", "Region", "Country",
    ],
    "silvApproach": [
        "Clear cut", "Close-to-nature silviculture",
        "Closer-to-nature silviculture", "Continuous cover forestry",
        "Agroforestry", "No management", "Other",
    ],
    "silvType": [
        "Tending and thinning", "Natural regeneration", "Plantation",
        "Uneven-aged with broadleaves", "Uneven-aged with conifers",
        "Transformation to uneven-aged", "Coppice", "Coppice under standards",
        "Forest pasture", "No intervention", "Other",
    ],
    "harvestSys": [
        "Ground-based logging", "Cable logging",
        "Helicopter logging", "Animal logging", "Other",
    ],
    "harvestType": ["Mainly manual", "Mainly by machine", "Mixed"],
    "harvestOp": ["Felling", "Processing", "Extraction"],
    "infra": ["Road network", "Skid trails", "Processing yards", "Other"],
    "cicesSection": ["Provisioning", "Regulation & maintenance", "Cultural"],
    "cicesDivision": [
        "Biomass provisioning",
        "Genetic materials from all biota",
        "Transformation of biochemical or physical inputs to ecosystems",
        "Regulation of baseline flows and extreme events",
        "Regulation of physical chemical biological conditions",
        "Physical and experiential interactions with natural environment",
        "Intellectual and representative interactions with natural environment",
        "Spiritual symbolic and other cultural interactions with natural environment",
    ],
    "cicesGroup": [
        "Cultivated terrestrial plants for nutrition materials or energy",
        "Wild plants for nutrition materials or energy",
        "Wild animals for nutrition materials or energy",
        "Genetic material from plants algae or fungi",
        "Reduction of nutrient loads and mediation of wastes or toxic substances",
        "Mediation of nuisances of anthropogenic origin",
        "Erosion control",
        "Hydrological cycle and water flow regulation",
        "Hazard mitigation",
        "Lifecycle maintenance habitat and gene pool protection",
        "Pest and disease control",
        "Regulation of soil quality",
        "Water conditions",
        "Atmospheric composition and conditions",
        "Direct in-situ and outdoor interactions with living systems",
        "Elements of living systems indirectly appreciated",
    ],
    "esInput": ["Biophysical", "Socioeconomic", "Expert-based", "Combined"],
    "esQuant": ["Biophysical", "Monetary", "Index-based", "Qualitative"],
    "esIndicator": ["Pressure", "State", "Benefit", "Value"],
    "ppTempScale": [
        "Long term (strategic)", "Medium term (tactical)",
        "Short term (operational)",
    ],
    "ppSpatScale": [
        "Stand level", "Forest level", "Regional level", "National level",
        "Continental level", "Global level",
    ],
    "ppDecision": [
        "Unilateral (one DM)", "Participatory (many DM no consensus)",
        "Collegial (many DM consensus needed)",
        "Single objective", "Multiple objectives",
    ],
    "ppSpatCtx": [
        "Spatial with neighbourhood interrelations",
        "Spatial with no neighbourhood interrelations",
        "Non-spatial",
    ],
    "csExtent": ["Local", "Regional", "National", "Global"],
    "csBiome": [
        "Alpine", "Anatolian", "Arctic", "Atlantic", "Boreal",
        "Continental", "Mediterranean", "Pannonian", "Steppic",
    ],
    "csClim": ["GWL 2C", "GWL 3C", "RCP 8.5", "SSP4", "Other"],
    "csLULC": ["LULUCF", "AFOLU", "Other"],
    "dataSource": [
        "Remote sensing", "Field survey", "Administrative",
        "Modelling", "Literature",
    ],
    "dataType": [
        "Vector", "Raster", "Database", "Table",
        "Time series", "Document", "Indicator",
    ],
    "spatScale": [
        "Tree", "Stand", "Forest", "Landscape",
        "Region", "National", "Global",
    ],
    "tempScale": [
        "Sub-annual", "Annual", "Decadal",
        "Event-based", "One-time", "Continuous",
    ],
    "outViz": [
        "Raster", "Vector", "Table", "Report", "Graph", "Time series",
        "Map", "Chart", "Text summary", "Interactive tool",
    ],
    "outTempScale": ["Sub-annual", "Annual", "Decadal", "Multi-decadal"],
    "outQuant": ["Biophysical", "Monetary", "Index-based", "Qualitative"],
    "outDataType": ["Quantitative", "Qualitative"],
    "madm": [
        "AHP", "ANP", "PROMETHEE", "ELECTRE", "Scoring",
        "Planning Balance Sheet", "Cost Benefit Analysis", "Other",
    ],
    "modm": ["Linear programming", "Integer programming", "Other"],
    "uncMethod": [
        "Monte Carlo Simulation", "Markov Chains", "Bayes Belief Networks",
        "Fuzzy sets", "Statistics", "Other",
    ],
    "uncSource": [
        "Input data", "Scenario assumptions", "Model structure",
        "Model parameterization", "Output evaluation",
    ],
    "scenType": [
        "Climate scenario", "Land use scenario", "Management scenario",
        "Socio-economic scenario", "Disturbance scenario",
        "Policy scenario", "Other",
    ],
    "engMode": [
        "Workshop", "Co-design", "Focus group",
        "Survey", "Stakeholder panel",
    ],
    "facilitation": ["Internal", "External", "Not facilitated"],
    "tempStruct": ["One-time", "Iterative", "Continuous"],
    "kmProcess": [
        "Knowledge Generation", "Knowledge Identification",
        "Knowledge Evaluation/Assessment", "Knowledge Storage",
        "Knowledge Application", "Knowledge Transfer",
    ],
    "kmTech": [
        "Artificial Intelligence", "Expert System", "Database",
        "Lessons Learned", "Best Practices",
        "Community of Practice/Webportal", "Knowledge Map/Cognitive Map",
    ],
    "actorRole": [
        "Developer", "Researcher", "Decision maker",
        "Provider", "Beneficiary", "Other",
    ],
    "actorInteraction": [
        "Top-down", "Participatory", "Co-production",
        "Contract-based", "Informal", "Hybrid",
    ],
    "actorLevel": [
        "Local", "Regional", "National", "International (regional)",
        "Supranational", "Global",
    ],
    "actorSector": [
        "Public", "Private", "Academic/Research", "NGO", "Community", "Mixed",
    ],
    "actorAuthority": [
        "Regulatory", "Advisory", "Operational", "Technical",
        "Financial", "Mixed", "None",
    ],
    # supplementary (needed by specific fields)
    "accessibility": ["Open", "Licensed", "Internal/restricted"],
    "modelClass": [
        "Model of vegetation dynamics",
        "Ecological model \u2013 Habitat model",
        "Ecological model \u2013 Carbon dynamics model",
        "Ecological model \u2013 Wildlife model",
        "Ecological model \u2013 Soil processes model",
        "Natural disturbance model",
        "Socio-economic model \u2013 Economic model",
        "Socio-economic model \u2013 Productivity model",
        "Socio-economic model \u2013 Non-wood forest product model",
    ],
    "esSpatRes": ["Local", "Regional", "National", "Global"],
    "esTempRes": ["One-time", "Seasonal", "Annual", "Long-term"],
}

# ── field type shorthands ─────────────────────────────────────────────────
T  = "text"
N  = "number"
U  = "url"
E  = "email"
TA = "textarea"
MI = "multi"
B3 = "bool3"


def S(v: str) -> str:
    return f"sel:{v}"


# ── sheet field definitions ───────────────────────────────────────────────
SHEET_FIELDS: dict[str, list[tuple[str, str, str]]] = {
    "DSS": [
        ("DSS_ID",                        "DSS Identifier",                   T),
        ("hasName",                       "Full Name",                        T),
        ("hasAcronym",                    "Acronym",                          T),
        ("hasDescription",                "Description",                      TA),
        ("hasRelatedDSS",                 "Related DSS",                      T),
        ("hasURL",                        "Website",                          U),
        ("hasOnlineDemo",                 "Online Demo URL",                  U),
        ("hasUserManual",                 "User Manual Exists",               B3),
        ("hasUserManualURL",              "User Manual URL",                  U),
        ("hasTechnicalDocumentation",     "Technical Documentation Exists",   B3),
        ("hasTechnicalDocumentationURL",  "Technical Documentation URL",      U),
        ("hasScientificReference",        "Scientific Reference",             T),
        ("hasAccessibility",              "Accessibility",                    S("accessibility")),
        ("hasCosts",                      "Ready-to-use Costs",               S("costs")),
        ("supportsProblemStructuring",    "Problem Structuring",              B3),
        ("supportsProblemModeling",       "Problem Modelling",                B3),
        ("supportsProblemSolving",        "Problem Solving",                  B3),
        ("wikiContact",                   "Wiki Contact Person",              T),
        ("wikiEmail",                     "Wiki Contact Email",               E),
    ],
    "Software": [
        ("DSS_ID",                     "DSS Identifier",                      T),
        ("hasName",                    "Software Name",                       T),
        ("hasVersion",                 "Version",                             T),
        ("hasITToolType",              "IT Tool Type",                        S("itToolType")),
        ("hasTechnologyReadiness",     "Technology Readiness",                S("techReadiness")),
        ("hasAccessibilityURL",        "Accessibility URL",                   U),
        ("requiresOperatingSystem",    "Operating System",                    MI),
        ("hasProgrammingLanguage",     "Programming Language",                MI),
        ("hasDatabaseManagementSystem","Database Management System",          MI),
        ("hasGISIntegration",          "GIS Integration",                     MI),
        ("usesAI",                     "AI Integration",                      T),
        ("hasHardwareRequirements",    "Hardware & Software Requirements",    TA),
        ("targetUserGroup",            "Target User Group",                   MI),
        ("hasCosts",                   "Ready-to-use Costs",                  S("costs")),
        ("requiresSkills",             "Specific Skill Requirements",         T),
        ("requiresTrainingsDays",      "Training Time Required",              S("training")),
        ("hasResponsibleOrganization", "Responsible Organisation",            T),
        ("contactPerson",              "Contact Person",                      T),
        ("contactEmail",               "Contact Email",                       E),
        ("hasOrigin",                  "Origin",                              S("origin")),
        ("hasDeveloperCount",          "Developer Team Size",                 S("devTeam")),
        ("devActors",                  "Actors Involved in Development",      MI),
        ("hasDesignMethodology",       "Design Methodology",                  T),
        ("hasDevelopmentManagement",   "Development Management",              S("devMgmt")),
        ("hasRealLifeApplication",     "Real-life Application Count",         S("realLife")),
        ("hasApplicationYear",         "Last Year of Real-life Application",  N),
        ("hasEducationalApplication",  "Utilisation in Education",            S("eduUse")),
        ("hasEduYear",                 "Last Year of Educational Use",        N),
        ("hasRDApplication",           "Utilisation in R&D",                  T),
        ("hasUpdateYear",              "Last Update Year",                    N),
        ("hasUpgradeYear",             "Last Upgrade Year",                   N),
        ("hasITDocumentation",         "IT/Code Documentation Exists",        B3),
        ("hasITDocumentationURL",      "IT Documentation URL",                U),
    ],
    "Model": [
        ("DSS_ID",               "DSS Identifier",                    T),
        ("hasName",              "Model Name",                        T),
        ("modelClass",           "Model Class (ontology subtype)",    S("modelClass")),
        ("isClimateSensitive",   "Climate Sensitive",                 B3),
        ("hasGrowth",            "Growth Submodel",                   B3),
        ("hasRegeneration",      "Regeneration Submodel",             B3),
        ("hasMortality",         "Mortality Submodel",                B3),
        ("hasBiome",             "Biome",                             MI),
        ("hasModelingType",      "Model Type",                        S("vegType")),
        ("hasModelName",         "Model Name (specific)",             T),
        ("spatScaleGrowth",      "Spatial Scale \u2013 Growth",       S("spatVeg")),
        ("spatScaleRegen",       "Spatial Scale \u2013 Regeneration", S("spatVeg")),
        ("spatScaleMort",        "Spatial Scale \u2013 Mortality",    S("spatVeg")),
        ("hasTemporalResolution","Temporal Resolution",               S("tempVeg")),
        ("hasDisturbanceModel",  "Has Disturbance Model",             B3),
        ("disturbanceTypes",     "Disturbance Types Represented",     MI),
        ("fireEmergent",         "Fire \u2013 Emergent",              B3),
        ("fireDeterministic",    "Fire \u2013 Deterministic",         B3),
        ("fireSpatial",          "Fire \u2013 Spatially Explicit",    B3),
        ("windEmergent",         "Wind \u2013 Emergent",              B3),
        ("windDeterministic",    "Wind \u2013 Deterministic",         B3),
        ("windSpatial",          "Wind \u2013 Spatially Explicit",    B3),
        ("beetleEmergent",       "Beetle \u2013 Emergent",            B3),
        ("beetleDeterministic",  "Beetle \u2013 Deterministic",       B3),
        ("beetleSpatial",        "Beetle \u2013 Spatially Explicit",  B3),
        ("snowEmergent",         "Snow/Ice \u2013 Emergent",          B3),
        ("snowDeterministic",    "Snow/Ice \u2013 Deterministic",     B3),
        ("snowSpatial",          "Snow/Ice \u2013 Spatially Explicit",B3),
        ("droughtEmergent",      "Drought \u2013 Emergent",           B3),
        ("droughtDeterministic", "Drought \u2013 Deterministic",      B3),
        ("droughtSpatial",       "Drought \u2013 Spatially Explicit", B3),
        ("bioticEmergent",       "Biotic (other) \u2013 Emergent",    B3),
        ("bioticDeterministic",  "Biotic (other) \u2013 Deterministic",B3),
        ("bioticSpatial",        "Biotic (other) \u2013 Spatially Explicit",B3),
        ("abioticEmergent",      "Abiotic (other) \u2013 Emergent",   B3),
        ("abioticDeterministic", "Abiotic (other) \u2013 Deterministic",B3),
        ("abioticSpatial",       "Abiotic (other) \u2013 Spatially Explicit",B3),
        ("soilNutrient",         "Soil \u2013 Nutrient Cycling",      B3),
        ("soilWater",            "Soil \u2013 Water Balance",         B3),
        ("wildlifePopDyn",       "Wildlife \u2013 Population Dynamics",B3),
        ("wildlifeSpecies",      "Wildlife \u2013 Species Represented",MI),
        ("nwfpModel",            "NWFP Model Present",                B3),
        ("nwfpTypes",            "NWFP Types",                        MI),
        ("habitatModel",         "Habitat Model Present",             B3),
        ("habitatSpecies",       "Habitat Model \u2013 Species Focus", MI),
        ("productivityModel",    "Productivity Model Present",        B3),
        ("carbonModel",          "Carbon Dynamics Model Present",     B3),
        ("hasTreeSpecies",       "Main Tree Species",                 MI),
        ("hasMixture",           "Mixed Forest Types",                B3),
        ("hasScientificReference","Scientific Reference",             T),
        ("hasURL",               "Model URL / Repository",            U),
    ],
    "ForestManagement": [
        ("DSS_ID",           "DSS Identifier",                T),
        ("hasName",          "Strategy / Configuration Name", T),
        ("hasRegion",        "Regional Area",                 MI),
        ("hasManagementScale","Management Scale",             MI),
        ("silvApproach",     "Silvicultural Approach",        MI),
        ("silvType",         "Silvicultural Type",            MI),
        ("mainTreeSpecies",  "Main Tree Species",             MI),
        ("hasMixture",       "Mixed Forest Types",            B3),
        ("biodivFeature",    "Biodiversity Features",         MI),
        ("harvestSystem",    "Harvesting System Type",        MI),
        ("harvestType",      "Harvesting Type",               MI),
        ("harvestOp",        "Harvesting Operation",          MI),
        ("infraMgmt",        "Infrastructure Management",     MI),
        ("hasDescription",   "Notes",                         TA),
    ],
    "EcosystemServices": [
        ("DSS_ID",             "DSS Identifier",                 T),
        ("hasName",            "ES Name / Label",                T),
        ("isModelled",         "Modelled in DSS",                B3),
        ("cicesSection",       "CICES Section",                  S("cicesSection")),
        ("cicesDivision",      "CICES Division",                 S("cicesDivision")),
        ("cicesGroup",         "CICES Group",                    S("cicesGroup")),
        ("cicesClass",         "CICES Class (most specific)",    T),
        ("modellingMethod",    "Modelling Method",               T),
        ("esInputType",        "ES Input Data Type",             MI),
        ("quantMethod",        "Quantification Method",          MI),
        ("indicatorSubtype",   "Indicator Subtype",              MI),
        ("hasUnit",            "Unit of Measurement",            T),
        ("outDataType",        "Output Data Type",               S("outDataType")),
        ("spatialResolution",  "Spatial Resolution",             S("esSpatRes")),
        ("temporalResolution", "Temporal Resolution",            S("esTempRes")),
        ("hasDescription",     "Notes",                          TA),
    ],
    "PlanningProblem": [
        ("DSS_ID",             "DSS Identifier",                 T),
        ("hasName",            "Planning Problem Title",         T),
        ("hasDescription",     "Description",                    TA),
        ("tempScale",          "Temporal Scale",                 MI),
        ("spatScale",          "Spatial Scale",                  MI),
        ("decisionSituation",  "Decision-Making Situation",      MI),
        ("spatContext",        "Spatial Context",                S("ppSpatCtx")),
        ("esDimension",        "Ecosystem Services Dimension",   T),
    ],
    "CaseStudy": [
        ("DSS_ID",                "DSS Identifier",                     T),
        ("hasName",               "Full Name",                          T),
        ("hasLocation",           "Location",                           T),
        ("hasCountry",            "Country",                            T),
        ("hasBiome",              "Biome / Eco-region",                 MI),
        ("geoExtent",             "Geographic Extent",                  S("csExtent")),
        ("tempCovStart",          "Temporal Coverage Start (year)",     N),
        ("tempCovEnd",            "Temporal Coverage End (year)",       N),
        ("duration",              "Duration (years)",                   N),
        ("climScenario",          "Climate Change Scenario",            MI),
        ("lulcTypology",          "Land Use Change Typology",           MI),
        ("naturalDisturbances",   "Assesses Natural Disturbances",      B3),
        ("sensitivityAnalysis",   "Sensitivity Analysis of Objectives", B3),
        ("hasIFM",                "Integrated Forest Management (IFM)", B3),
        ("hasSFM",                "Sustainable Forest Management (SFM)",B3),
        ("hasAFM",                "Adaptive Forest Management (AFM)",   B3),
        ("realLifeApp",           "Real-life Application",              B3),
        ("rdApp",                 "R&D Application",                    B3),
        ("esBeneficiary",         "Beneficiary of ES",                  MI),
        ("esProvider",            "Provider of ES",                     MI),
        ("hasScientificReference","Scientific Reference",               T),
        ("relatedPP",             "Related Planning Problem",           T),
        ("fullDescription",       "Full Description",                   TA),
    ],
    "InputData": [
        ("DSS_ID",              "DSS Identifier",             T),
        ("hasName",             "Dataset / Data Source Name", T),
        ("hasDataSource",       "Data Source Type",           MI),
        ("hasDataType",         "Data Type",                  MI),
        ("hasFormat",           "File Format",                MI),
        ("hasCollectionMethod", "Collection Method",          T),
        ("esInputType",         "ES Input Data Type",         MI),
        ("spatScale",           "Spatial Scale",              S("spatScale")),
        ("spatCoverage",        "Spatial Coverage",           T),
        ("tempScale",           "Temporal Scale",             S("tempScale")),
        ("tempCoverage",        "Temporal Coverage",          T),
        ("hasUnit",             "Unit",                       T),
        ("hasProvider",         "Provider",                   T),
        ("hasURL",              "URL",                        U),
        ("hasDescription",      "Notes",                      TA),
    ],
    "OutputData": [
        ("DSS_ID",              "DSS Identifier",                   T),
        ("hasName",             "Output Name",                      T),
        ("hasVisualization",    "Visualization Type",               MI),
        ("hasDataType",         "Output Data Type",                 S("outDataType")),
        ("hasFormat",           "File Format",                      MI),
        ("quantMethod",         "Quantification Method",            MI),
        ("dataProcessing",      "Data Processing Method",           MI),
        ("spatScale",           "Spatial Scale",                    S("spatScale")),
        ("tempScale",           "Temporal Scale",                   S("outTempScale")),
        ("hasUnit",             "Unit",                             T),
        ("uncertaintyIncluded", "Uncertainty Included in Output",   B3),
        ("hasDescription",      "Notes",                            TA),
    ],
    "DecisionTechniques": [
        ("DSS_ID",         "DSS Identifier",                       T),
        ("structuring",    "Problem Identification/Structuring",   B3),
        ("modelling",      "Problem Modelling",                    B3),
        ("solving",        "Problem Solving",                      B3),
        ("madmMethod",     "MADM Method(s)",                       MI),
        ("modmMethod",     "MODM Method(s)",                       MI),
        ("hasDescription", "Notes",                                TA),
    ],
    "UncertaintyEvaluation": [
        ("DSS_ID",         "DSS Identifier",                  T),
        ("modelsAssess",   "Models Assess Uncertainty",       B3),
        ("uncMethod",      "Type of Uncertainty Evaluation",  MI),
        ("uncSource",      "Uncertainty Source",              MI),
        ("hasDescription", "Notes",                           TA),
    ],
    "Scenario": [
        ("DSS_ID",         "DSS Identifier",        T),
        ("hasName",        "Scenario Name",         T),
        ("scenType",       "Scenario Type",         MI),
        ("tempStart",      "Time Horizon Start",    N),
        ("tempEnd",        "Time Horizon End",      N),
        ("hasDescription", "Scenario Description",  TA),
    ],
    "ParticipatoryProcess": [
        ("DSS_ID",           "DSS Identifier",            T),
        ("hasName",          "Process Name / Label",      T),
        ("modeOfEngagement", "Mode of Engagement",        MI),
        ("engPurpose",       "Engagement Purpose",        TA),
        ("facilitation",     "Facilitation",              S("facilitation")),
        ("toolsOrMethods",   "Tools or Methods Used",     T),
        ("tempStructure",    "Temporal Structure",        S("tempStruct")),
        ("evalMethod",       "Evaluation Method",         T),
        ("participantCount", "Participant Count",         N),
        ("hasDescription",   "Notes",                     TA),
    ],
    "LessonsLearned": [
        ("DSS_ID",            "DSS Identifier",                    T),
        ("hasTitle",          "Title",                             T),
        ("hasStatement",      "What? (statement)",                 TA),
        ("hasJustification",  "Why? (justification / evidence)",   TA),
        ("hasRecommendation", "How? (recommendation)",             TA),
        ("relatedDSS",        "Related DSS",                       T),
        ("relatedCase",       "Related Case Study",                T),
        ("relatedPP",         "Related Planning Problem",          T),
    ],
    "KMProcesses": [
        ("DSS_ID",        "DSS Identifier",               T),
        ("hasName",       "Label / Name",                 T),
        ("kmProcess",     "Supported KM Process",         MI),
        ("kmTechnique",   "Type of KM Techniques Used",   MI),
        ("hasDescription","Notes",                        TA),
    ],
    "Actor": [
        ("DSS_ID",           "DSS Identifier",            T),
        ("hasName",          "Full Name",                 T),
        ("hasEmail",         "Email",                     E),
        ("hasAffiliation",   "Affiliation",               T),
        ("hasURL",           "Profile URL",               U),
        ("actorRole",        "Role of Actor",             MI),
        ("actorInteraction", "Interaction Mode",          S("actorInteraction")),
        ("hasCountry",       "Country / Region",          T),
        ("actorLevel",       "Institutional Level",       S("actorLevel")),
        ("actorSector",      "Sector",                    S("actorSector")),
        ("actorAuthority",   "Authority Type",            S("actorAuthority")),
        ("orgSector",        "Sector of Organisation",    S("actorSector")),
    ],
}

SHEET_ORDER = [
    "DSS", "Software", "Model", "ForestManagement", "EcosystemServices",
    "PlanningProblem", "CaseStudy", "InputData", "OutputData",
    "DecisionTechniques", "UncertaintyEvaluation", "Scenario",
    "ParticipatoryProcess", "LessonsLearned", "KMProcesses", "Actor",
]

SHEET_DESCRIPTIONS = {
    "DSS":                   "Core DSS identification, documentation links, and decision support function",
    "Software":              "Software implementation, technology readiness, development history",
    "Model":                 "Computational models: vegetation dynamics, disturbance, wildlife, carbon, etc.",
    "ForestManagement":      "Silvicultural approaches, tree species, harvesting systems",
    "EcosystemServices":     "ES addressed by the DSS using the CICES classification hierarchy",
    "PlanningProblem":       "Decision-making context, temporal/spatial scale, objectives",
    "CaseStudy":             "Empirical DSS applications in real or study contexts",
    "InputData":             "Data sources and datasets used as DSS inputs",
    "OutputData":            "Results and indicators produced by the DSS",
    "DecisionTechniques":    "MADM/MODM analytical methods used in the DSS",
    "UncertaintyEvaluation": "How the DSS handles and communicates uncertainty",
    "Scenario":              "Future conditions or alternative management states",
    "ParticipatoryProcess":  "Stakeholder involvement in DSS design or application",
    "LessonsLearned":        "Insights and recommendations from DSS applications",
    "KMProcesses":           "Knowledge management activities supported by the DSS",
    "Actor":                 "People involved as developers, researchers, decision-makers, or stakeholders",
}

# ── field metadata: (ontology definition, default Heureka example) ────────
# All examples refer to the Heureka DSS (Swedish University of Agricultural
# Sciences) unless overridden per sheet in FIELD_EXAMPLES below.
FIELD_META: dict[str, tuple[str, str]] = {
    # universal
    "DSS_ID": (
        "Unique identifier shared across all sheets to link all instances to the same DSS. Use the DSS acronym or a short slug.",
        "Heureka",
    ),
    # DSS
    "hasName": (
        "The full official name of the entity (DSS, software, model, actor, etc.).",
        "Heureka Decision Support System",
    ),
    "hasAcronym": (
        "A short abbreviation by which the DSS is commonly known in literature or practice.",
        "Heureka",
    ),
    "hasDescription": (
        "Free-text description providing context about the instance.",
        "Multi-objective forest planning system supporting long-term strategic decisions across Swedish boreal forests.",
    ),
    "hasRelatedDSS": (
        "Name of another DSS that is related, complementary, or a predecessor to this one.",
        "SILVA-DSS",
    ),
    "hasURL": (
        "Web address of the official homepage, repository, or access page.",
        "https://www.heurekaslu.se/",
    ),
    "hasOnlineDemo": (
        "URL of an online demonstration or interactive example of the DSS.",
        "https://www.heurekaslu.se/wiki/index.php/Demo",
    ),
    "hasUserManual": (
        "Whether a user manual or end-user guide exists for this DSS.",
        "Yes",
    ),
    "hasUserManualURL": (
        "URL pointing to the user manual or end-user documentation.",
        "https://www.heurekaslu.se/wiki/index.php/Heureka_Wiki",
    ),
    "hasTechnicalDocumentation": (
        "Whether technical or architectural documentation exists beyond the user manual.",
        "Yes",
    ),
    "hasTechnicalDocumentationURL": (
        "URL pointing to the technical documentation.",
        "https://www.heurekaslu.se/wiki/index.php/Technical_documentation",
    ),
    "hasScientificReference": (
        "Bibliographic reference (Author, Year, Journal/Book) for the key publication describing this entity.",
        "Wikström et al. (2011) Scandinavian Journal of Forest Research 26(S10): 85\u201396",
    ),
    "hasAccessibility": (
        "The licence or access model under which the DSS is made available to users.",
        "Open",
    ),
    "hasCosts": (
        "Approximate cost for an end user to obtain and use the DSS in its ready-to-use form.",
        "0 \u20ac",
    ),
    "supportsProblemStructuring": (
        "Whether the DSS provides tools to help users identify and articulate the decision problem, objectives, criteria, and stakeholders.",
        "Yes",
    ),
    "supportsProblemModeling": (
        "Whether the DSS formally represents relationships between decision elements (criteria, alternatives, constraints).",
        "Yes",
    ),
    "supportsProblemSolving": (
        "Whether the DSS generates, evaluates, or selects among alternative management solutions.",
        "Yes",
    ),
    "wikiContact": (
        "Name of the person responsible for maintaining or updating the DSS entry in the ForestDSS wiki.",
        "Peder Wikstr\u00f6m",
    ),
    "wikiEmail": (
        "Email address of the wiki contact person.",
        "peder.wikstrom@slu.se",
    ),
    # Software
    "hasVersion": (
        "Version number or identifier of the software at the time of characterization.",
        "2.14",
    ),
    "hasITToolType": (
        "Category of software interface through which users interact with the DSS.",
        "GUI \u2013 desktop",
    ),
    "hasTechnologyReadiness": (
        "Maturity level of the software, from early prototype to commercial product.",
        "Non-commercial product",
    ),
    "hasAccessibilityURL": (
        "URL linking to the download, access, or registration page for the software.",
        "https://www.heurekaslu.se/download",
    ),
    "requiresOperatingSystem": (
        "Operating system(s) required or supported. Enter multiple values separated by semicolons.",
        "Windows",
    ),
    "hasProgrammingLanguage": (
        "Programming language(s) used to implement the software. Enter multiple values separated by semicolons.",
        "C#",
    ),
    "hasDatabaseManagementSystem": (
        "Database management system(s) used internally by the software. Enter multiple values separated by semicolons.",
        "SQLite",
    ),
    "hasGISIntegration": (
        "GIS software the DSS integrates with or depends on. Enter multiple values separated by semicolons.",
        "ESRI (ArcGIS)",
    ),
    "usesAI": (
        "Description of any artificial intelligence or machine learning components used in the software.",
        "No AI components",
    ),
    "hasHardwareRequirements": (
        "Minimum or recommended hardware and software dependencies beyond the listed OS.",
        "Windows 10 or later; 4 GB RAM; .NET Framework 4.8",
    ),
    "targetUserGroup": (
        "Intended primary user groups. Enter multiple values separated by semicolons.",
        "Forest service/administration; Research institution; Forest owner/manager",
    ),
    "requiresSkills": (
        "Specific knowledge or technical skills required to operate the DSS effectively.",
        "Basic forest management knowledge; Windows OS familiarity",
    ),
    "requiresTrainingsDays": (
        "Approximate time needed for a new user to become proficient in the DSS.",
        "\u2264 1 week",
    ),
    "hasResponsibleOrganization": (
        "Organisation legally or operationally responsible for the DSS.",
        "Swedish University of Agricultural Sciences (SLU)",
    ),
    "contactPerson": (
        "Person to contact regarding the software (may differ from the wiki contact).",
        "Peder Wikstr\u00f6m",
    ),
    "contactEmail": (
        "Email address for the software contact person.",
        "peder.wikstrom@slu.se",
    ),
    "hasOrigin": (
        "Institutional context in which the DSS was originally developed.",
        "Research institution",
    ),
    "hasDeveloperCount": (
        "Approximate number of people with at least 10% involvement in software development.",
        "> 5",
    ),
    "devActors": (
        "Categories of actors involved in development. Enter multiple values separated by semicolons.",
        "Researchers; Forest specialists",
    ),
    "hasDesignMethodology": (
        "Design or development methodology used (e.g. user-centred design, agile, participatory design).",
        "User-centred design",
    ),
    "hasDevelopmentManagement": (
        "Software project management approach used during development.",
        "None",
    ),
    "hasRealLifeApplication": (
        "Approximate total number of real operational applications of the DSS reported.",
        "> 100",
    ),
    "hasApplicationYear": (
        "Most recent year in which the DSS was used in a real operational context.",
        "2024",
    ),
    "hasEducationalApplication": (
        "Whether and how the DSS is used in an educational context.",
        "Yes \u2013 direct use by students",
    ),
    "hasEduYear": (
        "Most recent year in which the DSS was used in education.",
        "2024",
    ),
    "hasRDApplication": (
        "Description or URL of known uses of the DSS in research and development.",
        "Ongoing use in SLU research on climate adaptation of Swedish forests",
    ),
    "hasUpdateYear": (
        "Year of the most recent minor update or maintenance release.",
        "2024",
    ),
    "hasUpgradeYear": (
        "Year of the most recent major version upgrade with significant new functionality.",
        "2023",
    ),
    "hasITDocumentation": (
        "Whether code-level or API documentation exists for the software.",
        "Yes",
    ),
    "hasITDocumentationURL": (
        "URL pointing to IT or code documentation (e.g. GitHub, developer wiki).",
        "https://github.com/HeurekaSLU",
    ),
    # Model
    "modelClass": (
        "Most specific ontology subclass that best describes this model's primary function.",
        "Model of vegetation dynamics",
    ),
    "isClimateSensitive": (
        "Whether at least one modelled process (growth, mortality, regeneration) explicitly responds to temperature or precipitation inputs.",
        "Yes",
    ),
    "hasGrowth": (
        "Whether the model includes a growth submodel for individual trees or cohorts.",
        "Yes",
    ),
    "hasRegeneration": (
        "Whether the model includes a regeneration submodel for seedling establishment and early development.",
        "Yes",
    ),
    "hasMortality": (
        "Whether the model includes a mortality submodel representing natural tree death.",
        "Yes",
    ),
    "hasBiome": (
        "Forest biome(s) or vegetation zones for which the model was designed or validated. Enter multiple values separated by semicolons.",
        "Boreal; Temperate continental",
    ),
    "hasModelingType": (
        "Fundamental modelling approach: Empirical (statistical functions fitted to data), Process-based (mechanistic), or Hybrid.",
        "Empirical",
    ),
    "hasModelName": (
        "Specific name of the model or model family (e.g. ForClim, SORTIE, iLand, 3-PG).",
        "Elfving growth functions",
    ),
    "spatScaleGrowth": (
        "Spatial unit at which the growth submodel operates.",
        "Stand",
    ),
    "spatScaleRegen": (
        "Spatial unit at which the regeneration submodel operates.",
        "Stand",
    ),
    "spatScaleMort": (
        "Spatial unit at which the mortality submodel operates.",
        "Stand",
    ),
    "hasTemporalResolution": (
        "Time step at which the vegetation dynamics model advances the simulation.",
        "Annual",
    ),
    "hasDisturbanceModel": (
        "Whether the DSS includes an explicit natural disturbance model.",
        "Yes",
    ),
    "disturbanceTypes": (
        "Natural disturbance types represented in the model. Enter multiple values separated by semicolons.",
        "Wind; Fire",
    ),
    "fireEmergent": (
        "Fire arises from modelled interactions between fuel, weather, and ignition rather than being imposed as an external event.",
        "No",
    ),
    "fireDeterministic": (
        "Same inputs always produce the same fire outcome \u2014 no stochastic component.",
        "No",
    ),
    "fireSpatial": (
        "Fire spread is modelled across space, not only at stand level.",
        "No",
    ),
    "windEmergent": (
        "Wind damage arises from modelled interactions between stand structure, terrain, and wind field rather than an external trigger.",
        "No",
    ),
    "windDeterministic": (
        "Same inputs always produce the same wind damage outcome \u2014 no stochastic component.",
        "No",
    ),
    "windSpatial": (
        "Wind damage is modelled across space, not only at stand level.",
        "No",
    ),
    "beetleEmergent": (
        "Bark beetle outbreak arises from modelled host\u2013beetle\u2013climate interactions rather than an external trigger.",
        "Unknown",
    ),
    "beetleDeterministic": (
        "Same inputs always produce the same bark beetle outcome \u2014 no stochastic component.",
        "Unknown",
    ),
    "beetleSpatial": (
        "Bark beetle spread is modelled across space, not only at stand level.",
        "Unknown",
    ),
    "snowEmergent": (
        "Snow/ice damage arises from modelled interactions between stand structure, load, and weather rather than an external trigger.",
        "No",
    ),
    "snowDeterministic": (
        "Same inputs always produce the same snow/ice damage outcome \u2014 no stochastic component.",
        "No",
    ),
    "snowSpatial": (
        "Snow/ice damage is modelled across space, not only at stand level.",
        "No",
    ),
    "droughtEmergent": (
        "Drought stress arises from modelled soil water balance, evapotranspiration, and tree physiology rather than an external trigger.",
        "No",
    ),
    "droughtDeterministic": (
        "Same inputs always produce the same drought outcome \u2014 no stochastic component.",
        "No",
    ),
    "droughtSpatial": (
        "Drought effects are modelled across space, not only at stand level.",
        "No",
    ),
    "bioticEmergent": (
        "Biotic disturbance (other than bark beetle) arises from modelled ecological interactions rather than an external trigger.",
        "Unknown",
    ),
    "bioticDeterministic": (
        "Same inputs always produce the same biotic disturbance outcome \u2014 no stochastic component.",
        "Unknown",
    ),
    "bioticSpatial": (
        "Biotic disturbance is modelled across space, not only at stand level.",
        "Unknown",
    ),
    "abioticEmergent": (
        "Abiotic disturbance (other than fire, wind, snow, drought) arises from modelled physical processes rather than an external trigger.",
        "Unknown",
    ),
    "abioticDeterministic": (
        "Same inputs always produce the same abiotic disturbance outcome \u2014 no stochastic component.",
        "Unknown",
    ),
    "abioticSpatial": (
        "Abiotic disturbance effects are modelled across space, not only at stand level.",
        "Unknown",
    ),
    "soilNutrient": (
        "Whether the model simulates nutrient cycling dynamics in the soil.",
        "Yes",
    ),
    "soilWater": (
        "Whether the model simulates water balance processes in the soil.",
        "Yes",
    ),
    "wildlifePopDyn": (
        "Whether the model includes a wildlife population dynamics submodel.",
        "No",
    ),
    "wildlifeSpecies": (
        "Wildlife or biodiversity species groups represented in the model. Enter multiple values separated by semicolons.",
        "Bird species; Insects",
    ),
    "nwfpModel": (
        "Whether the model includes a non-wood forest product (NWFP) submodel.",
        "No",
    ),
    "nwfpTypes": (
        "Types of non-wood forest products modelled. Enter multiple values separated by semicolons.",
        "Berries; Mushroom",
    ),
    "habitatModel": (
        "Whether the model includes a habitat suitability or biodiversity habitat model.",
        "Yes",
    ),
    "habitatSpecies": (
        "Species or groups targeted by the habitat model. Enter multiple values separated by semicolons.",
        "Bird species; Lichens",
    ),
    "productivityModel": (
        "Whether the model includes a forest productivity or timber volume production submodel.",
        "Yes",
    ),
    "carbonModel": (
        "Whether the model includes a carbon dynamics submodel tracking carbon stocks and fluxes.",
        "Yes",
    ),
    "hasTreeSpecies": (
        "Main tree species or groups that the model simulates. Enter multiple values separated by semicolons.",
        "Spruce; Pine; Birch",
    ),
    "hasMixture": (
        "Whether the model can simulate mixed-species forest stands.",
        "Yes",
    ),
    # ForestManagement
    "hasRegion": (
        "Geographic region(s) for which this management strategy is designed or applicable. Enter multiple values separated by semicolons.",
        "Northern Europe",
    ),
    "hasManagementScale": (
        "Spatial scale(s) at which management decisions are made. Enter multiple values separated by semicolons.",
        "Stand; Forest management unit",
    ),
    "silvApproach": (
        "Overall silvicultural approach or forest management system applied. Enter multiple values separated by semicolons.",
        "Clear cut; Close-to-nature silviculture",
    ),
    "silvType": (
        "Specific silvicultural treatment type(s) applied. Enter multiple values separated by semicolons.",
        "Tending and thinning; Natural regeneration; Plantation",
    ),
    "mainTreeSpecies": (
        "Dominant or managed tree species in the stands. Enter multiple values separated by semicolons.",
        "Spruce; Pine; Birch",
    ),
    "biodivFeature": (
        "Biodiversity features explicitly retained or managed. Enter multiple values separated by semicolons.",
        "Deadwood; Old trees",
    ),
    "harvestSystem": (
        "Harvesting technology and extraction system used. Enter multiple values separated by semicolons.",
        "Ground-based logging",
    ),
    "harvestType": (
        "Degree of mechanisation of harvesting operations. Enter multiple values separated by semicolons.",
        "Mainly by machine",
    ),
    "harvestOp": (
        "Specific harvesting operations performed. Enter multiple values separated by semicolons.",
        "Felling; Processing; Extraction",
    ),
    "infraMgmt": (
        "Forest infrastructure elements managed as part of this strategy. Enter multiple values separated by semicolons.",
        "Road network; Skid trails",
    ),
    # EcosystemServices
    "isModelled": (
        "Whether this ecosystem service is explicitly represented or calculated by the DSS.",
        "Yes",
    ),
    "cicesSection": (
        "Top-level CICES classification section of the ecosystem service (Provisioning, Regulation & maintenance, or Cultural).",
        "Provisioning",
    ),
    "cicesDivision": (
        "CICES Division providing a more specific classification within the Section.",
        "Biomass provisioning",
    ),
    "cicesGroup": (
        "CICES Group providing a detailed classification within the Division.",
        "Cultivated terrestrial plants for nutrition materials or energy",
    ),
    "cicesClass": (
        "Most specific CICES Class. Enter the class name and code, e.g. 'Cultivated terrestrial plants grown for nutritional purposes (1.1.1.1)'.",
        "Cultivated terrestrial plants grown for nutritional purposes (1.1.1.1)",
    ),
    "modellingMethod": (
        "Method or model used to quantify this ecosystem service (short description or reference).",
        "Empirical growth functions combined with harvest scheduling optimisation",
    ),
    "esInputType": (
        "Type of input data used to assess this ES. Enter multiple values separated by semicolons.",
        "Biophysical",
    ),
    "quantMethod": (
        "Quantification method used to measure or express this ES. Enter multiple values separated by semicolons.",
        "Biophysical; Monetary",
    ),
    "indicatorSubtype": (
        "Indicator subtype(s) used. Pressure = human impact on ES. State = current ES condition. Benefit = what people receive. Value = monetary/non-monetary valuation. Enter multiple values separated by semicolons.",
        "State; Benefit",
    ),
    "hasUnit": (
        "Unit of measurement used to quantify the ES or data field.",
        "m\u00b3/ha/year",
    ),
    "outDataType": (
        "Whether output is expressed in quantitative (numeric) or qualitative (descriptive) terms.",
        "Quantitative",
    ),
    "spatialResolution": (
        "Spatial grain or extent at which the ES is estimated.",
        "Regional",
    ),
    "temporalResolution": (
        "Temporal grain or frequency at which the ES is estimated.",
        "Annual",
    ),
    # PlanningProblem
    "tempScale": (
        "Temporal planning horizon addressed by this problem or dataset. Enter multiple values separated by semicolons.",
        "Long term (strategic); Medium term (tactical)",
    ),
    "spatScale": (
        "Spatial scale of planning or data. Select from dropdown or enter multiple values separated by semicolons.",
        "Stand; Forest",
    ),
    "decisionSituation": (
        "Decision-making configuration: how many actors are involved, whether consensus is required, and how many objectives. Enter multiple values separated by semicolons.",
        "Multiple objectives; Unilateral (one DM)",
    ),
    "spatContext": (
        "Whether the planning problem involves spatial interactions between adjacent management units.",
        "Spatial with no neighbourhood interrelations",
    ),
    "esDimension": (
        "Names of the ecosystem services explicitly addressed in this planning problem.",
        "Timber production; Carbon storage; Recreation; Biodiversity",
    ),
    # CaseStudy
    "hasLocation": (
        "Geographic location of the case study area (municipality, region, or site name).",
        "Central Sweden",
    ),
    "hasCountry": (
        "Country or countries where the case study was conducted.",
        "Sweden",
    ),
    "hasBiome": (
        "Biogeographical region or ecological zone of the study area. Enter multiple values separated by semicolons.",
        "Boreal; Continental",
    ),
    "geoExtent": (
        "Overall geographic scale of the case study.",
        "National",
    ),
    "tempCovStart": (
        "First year of the case study's planning horizon or data coverage.",
        "2010",
    ),
    "tempCovEnd": (
        "Last year of the case study's planning horizon or data coverage.",
        "2060",
    ),
    "duration": (
        "Total length of the planning horizon or study period in years.",
        "50",
    ),
    "climScenario": (
        "Climate change scenario(s) used in the case study. Enter multiple values separated by semicolons.",
        "GWL 2C",
    ),
    "lulcTypology": (
        "Land use/land cover change accounting framework(s) applied. Enter multiple values separated by semicolons.",
        "LULUCF",
    ),
    "naturalDisturbances": (
        "Whether the case study explicitly assesses or models natural disturbances (fire, wind, drought, pests).",
        "Yes",
    ),
    "sensitivityAnalysis": (
        "Whether the case study tests sensitivity of outcomes to changes in management objectives or weights.",
        "Yes",
    ),
    "hasIFM": (
        "Whether Integrated Forest Management (explicitly balancing timber with biodiversity and other ES) is applied.",
        "Yes",
    ),
    "hasSFM": (
        "Whether Sustainable Forest Management meeting FSC/PEFC or equivalent criteria is applied.",
        "Yes",
    ),
    "hasAFM": (
        "Whether Adaptive Forest Management (iterative management incorporating monitoring feedback) is applied.",
        "No",
    ),
    "realLifeApp": (
        "Whether the DSS was applied to an actual operational management decision (not demonstration only).",
        "Yes",
    ),
    "rdApp": (
        "Whether the DSS was used in a research/development context to test methods or compare scenarios.",
        "Yes",
    ),
    "esBeneficiary": (
        "Who receives the benefits of the ecosystem services in this case study. Enter multiple values separated by semicolons.",
        "State-owned enterprise; Private landowners",
    ),
    "esProvider": (
        "Who manages or owns the forest land providing the ecosystem services. Enter multiple values separated by semicolons.",
        "State-owned enterprise; Private landowners",
    ),
    "relatedPP": (
        "Name of the Planning Problem instance associated with this case study or lesson.",
        "Long-term multi-objective forest planning",
    ),
    "fullDescription": (
        "Comprehensive free-text description of the case study, including context, objectives, and outcomes.",
        "Strategic forest planning across Sweden\u2019s boreal forests targeting timber and carbon trade-offs under climate change.",
    ),
    # InputData
    "hasDataSource": (
        "Type(s) of source from which input data were collected. Enter multiple values separated by semicolons.",
        "Field survey; Administrative",
    ),
    "hasDataType": (
        "Data structure or format type(s) of the dataset. Enter multiple values separated by semicolons.",
        "Database; Table",
    ),
    "hasFormat": (
        "File format(s) in which data are provided. Enter multiple values separated by semicolons.",
        "Excel; CSV",
    ),
    "hasCollectionMethod": (
        "Method used to collect or generate the input data.",
        "Systematic sampling with field measurements",
    ),
    "spatCoverage": (
        "Geographic description of the area covered by the dataset.",
        "Sweden (national coverage)",
    ),
    "tempCoverage": (
        "Time range covered by the dataset (e.g. 2000\u20132020).",
        "2000\u20132020",
    ),
    "hasProvider": (
        "Organisation or person responsible for providing or publishing the dataset.",
        "Swedish Forest Agency (Skogsstyrelsen)",
    ),
    # OutputData
    "hasVisualization": (
        "Format(s) in which outputs are displayed or communicated. Enter multiple values separated by semicolons.",
        "Table; Graph; Time series; Map",
    ),
    "dataProcessing": (
        "Analytical or computational method used to process or derive this output. Enter multiple values separated by semicolons.",
        "Statistics",
    ),
    "uncertaintyIncluded": (
        "Whether this output includes an explicit representation of uncertainty (e.g. confidence interval, scenario range).",
        "Yes",
    ),
    # DecisionTechniques
    "structuring": (
        "Whether the DSS includes tools for problem structuring: identifying objectives, criteria, constraints, and stakeholders.",
        "Yes",
    ),
    "modelling": (
        "Whether the DSS formally represents the decision problem structure and the relationships between elements.",
        "Yes",
    ),
    "solving": (
        "Whether the DSS generates or selects among alternative management solutions.",
        "Yes",
    ),
    "madmMethod": (
        "Multi-attribute decision-making method(s) used to evaluate alternatives against multiple criteria. Enter multiple values separated by semicolons.",
        "AHP; Scoring",
    ),
    "modmMethod": (
        "Multi-objective decision-making method(s) used to optimise management solutions. Enter multiple values separated by semicolons.",
        "Linear programming",
    ),
    # UncertaintyEvaluation
    "modelsAssess": (
        "Whether the DSS explicitly assesses or quantifies uncertainty in its outputs or model results.",
        "Yes",
    ),
    "uncMethod": (
        "Method(s) used to evaluate or represent uncertainty. Enter multiple values separated by semicolons.",
        "Statistics; Monte Carlo Simulation",
    ),
    "uncSource": (
        "Source(s) of uncertainty addressed by the evaluation. Enter multiple values separated by semicolons.",
        "Input data; Model parameterization",
    ),
    # Scenario
    "scenType": (
        "Type(s) of scenario used to define alternative future conditions. Enter multiple values separated by semicolons.",
        "Management scenario; Climate scenario",
    ),
    "tempStart": (
        "First year of the scenario's time horizon.",
        "2020",
    ),
    "tempEnd": (
        "Last year of the scenario's time horizon.",
        "2100",
    ),
    # ParticipatoryProcess
    "modeOfEngagement": (
        "Format(s) through which stakeholders were engaged. Enter multiple values separated by semicolons.",
        "Workshop; Focus group",
    ),
    "engPurpose": (
        "Specific description of what participants were asked to contribute or decide during the process.",
        "Validate ES weights for multi-criteria planning model and prioritise management scenarios.",
    ),
    "facilitation": (
        "Whether the process was facilitated by the research team (Internal), an independent facilitator (External), or not facilitated.",
        "Internal",
    ),
    "toolsOrMethods": (
        "Specific structured tools or methods used during the participatory process.",
        "Delphi method; Structured decision-making",
    ),
    "tempStructure": (
        "Whether the process occurred once, in repeated rounds, or continuously throughout the project.",
        "One-time",
    ),
    "evalMethod": (
        "How the quality or outcome of the participatory process was evaluated.",
        "Participant feedback forms",
    ),
    "participantCount": (
        "Total number of participants across all events in the participatory process.",
        "24",
    ),
    # LessonsLearned
    "hasTitle": (
        "Short descriptive title summarising the lesson learned.",
        "Spatial resolution matters for biodiversity ES quantification",
    ),
    "hasStatement": (
        "The core lesson itself \u2014 what was observed, found, or concluded.",
        "Stand-level inputs significantly improve biodiversity ES estimates compared to national averages.",
    ),
    "hasJustification": (
        "Evidence or reasoning supporting the lesson.",
        "Stand-level NFI data reduced biodiversity indicator estimation error by ~30% vs. national averages.",
    ),
    "hasRecommendation": (
        "Concrete recommendation for how this lesson should be applied in future DSS use or development.",
        "Collect stand-level inventory data when biodiversity ES are a primary planning objective.",
    ),
    "relatedDSS": (
        "DSS_ID of the DSS to which this lesson or actor reference applies.",
        "Heureka",
    ),
    "relatedCase": (
        "Name or ID of the case study from which this lesson was derived.",
        "Swedish NFI strategic planning application",
    ),
    # KMProcesses
    "kmProcess": (
        "Knowledge management process(es) supported by this activity or tool. Enter multiple values separated by semicolons.",
        "Knowledge Storage; Knowledge Transfer; Knowledge Application",
    ),
    "kmTechnique": (
        "Type(s) of KM technique or tool used. Enter multiple values separated by semicolons.",
        "Database; Community of Practice/Webportal",
    ),
    # Actor
    "hasEmail": (
        "Professional email address of the actor.",
        "peder.wikstrom@slu.se",
    ),
    "hasAffiliation": (
        "Organisation or institution the actor is affiliated with.",
        "Swedish University of Agricultural Sciences (SLU)",
    ),
    "actorRole": (
        "Role(s) of this actor in relation to the DSS or case study. Enter multiple values separated by semicolons.",
        "Developer; Researcher",
    ),
    "actorInteraction": (
        "Mode of interaction between this actor and the DSS or decision process.",
        "Top-down",
    ),
    "actorLevel": (
        "Institutional or governance level at which this actor primarily operates.",
        "National",
    ),
    "actorSector": (
        "Sector of the organisation this actor belongs to.",
        "Academic/Research",
    ),
    "actorAuthority": (
        "Type of decision-making authority held by this actor in the forest management context.",
        "Technical",
    ),
    "orgSector": (
        "Sector of the organisation the actor belongs to (may equal actorSector for individual actors).",
        "Academic/Research",
    ),
}

# ── per-sheet example overrides (sheet_name, field_key) → example ─────────
FIELD_EXAMPLES: dict[tuple[str, str], str] = {
    # hasName
    ("Software",            "hasName"): "HeurekaDSS",
    ("Model",               "hasName"): "Heureka forest growth simulator",
    ("ForestManagement",    "hasName"): "Swedish business-as-usual clearcut forestry",
    ("EcosystemServices",   "hasName"): "Timber production",
    ("PlanningProblem",     "hasName"): "Long-term multi-objective forest planning",
    ("CaseStudy",           "hasName"): "Swedish NFI strategic planning application",
    ("InputData",           "hasName"): "National forest inventory sample plots",
    ("OutputData",          "hasName"): "Standing volume projections 2020\u20132100",
    ("Scenario",            "hasName"): "Business-as-usual 2020\u20132100",
    ("ParticipatoryProcess","hasName"): "Forest stakeholder ES trade-off workshop",
    ("KMProcesses",         "hasName"): "Heureka Wiki knowledge base",
    ("Actor",               "hasName"): "Peder Wikstr\u00f6m",
    # hasDescription / scenario desc
    ("ForestManagement",    "hasDescription"): "Conventional clearcut management with retention trees for biodiversity.",
    ("EcosystemServices",   "hasDescription"): "Notes on carbon stock accounting methodology.",
    ("PlanningProblem",     "hasDescription"): "Multi-criteria optimisation of timber production and carbon storage across a forest holding.",
    ("InputData",           "hasDescription"): "Systematic sample plots with tree-level measurements used as DSS input.",
    ("OutputData",          "hasDescription"): "Annual standing volume projections by forest owner category.",
    ("DecisionTechniques",  "hasDescription"): "Heureka uses LP for harvest scheduling; AHP for ES weighting.",
    ("UncertaintyEvaluation","hasDescription"): "Monte Carlo simulation applied to growth model parameters.",
    ("Scenario",            "hasDescription"): "Reference scenario based on current Swedish practices under RCP 4.5.",
    ("ParticipatoryProcess","hasDescription"): "Single workshop with regional forest managers to validate ES trade-off weights.",
    ("KMProcesses",         "hasDescription"): "Heureka Wiki provides documented guidance and best practices for practitioners.",
    # hasURL
    ("DSS",                 "hasURL"): "https://www.heurekaslu.se/",
    ("Model",               "hasURL"): "https://www.heurekaslu.se/wiki/index.php/Growth_functions",
    ("InputData",           "hasURL"): "https://www.skogsstyrelsen.se/",
    ("Actor",               "hasURL"): "https://www.slu.se/cv/peder-wikstrom/",
    # hasScientificReference
    ("Model",               "hasScientificReference"): "Elfving B. (2010) Development of a new growth function. SLU, Ume\u00e5.",
    ("CaseStudy",           "hasScientificReference"): "Wikstr\u00f6m et al. (2011) Scand J Forest Res 26(S10): 85\u201396",
    # hasUnit
    ("InputData",           "hasUnit"): "trees/ha; m\u00b3/ha",
    ("OutputData",          "hasUnit"): "m\u00b3/ha",
    # tempScale / spatScale context differences
    ("InputData",           "tempScale"): "Annual",
    ("OutputData",          "tempScale"): "Annual",
    ("InputData",           "spatScale"): "Stand",
    ("OutputData",          "spatScale"): "Stand",
    # quantMethod context
    ("OutputData",          "quantMethod"): "Biophysical; Monetary",
    # hasDataType context (InputData vs OutputData)
    ("InputData",           "hasDataType"): "Database; Table",
    ("OutputData",          "hasDataType"): "Quantitative",
    # hasCountry (Actor vs CaseStudy)
    ("Actor",               "hasCountry"): "Sweden",
    ("CaseStudy",           "hasCountry"): "Sweden",
}


# ── helpers ───────────────────────────────────────────────────────────────

_YEAR_KEYS = frozenset({
    "hasApplicationYear", "hasEduYear", "hasUpdateYear", "hasUpgradeYear",
    "tempCovStart", "tempCovEnd", "tempStart", "tempEnd",
})


def _fill_instruction(ftype: str, key: str = "") -> str:
    """Return a short filling instruction based on field type."""
    if ftype == T:
        return "Free text"
    if ftype == N:
        return "Number (year, e.g. 2024)" if key in _YEAR_KEYS else "Number"
    if ftype == U:
        return "URL (e.g. https://...)"
    if ftype == E:
        return "Email address"
    if ftype == TA:
        return "Free text (multi-line)"
    if ftype == MI:
        return "Separate multiple values with semicolons"
    if ftype == B3:
        return "Select from dropdown: Yes / No / Unknown"
    if ftype.startswith("sel:"):
        return "Select from dropdown (controlled vocabulary)"
    return "Free text"


def _type_label(ftype: str) -> str:
    """Return a human-readable type label for the field reference table."""
    if ftype == T:
        return "Text"
    if ftype == N:
        return "Number"
    if ftype == U:
        return "URL"
    if ftype == E:
        return "Email"
    if ftype == TA:
        return "Text (multi-line)"
    if ftype == MI:
        return "Multi-value (semicolons)"
    if ftype == B3:
        return "Dropdown: Yes/No/Unknown"
    if ftype.startswith("sel:"):
        return "Dropdown (controlled)"
    return "Text"


# ── style helpers ─────────────────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _col_width(label: str, key: str, ftype: str) -> float:
    raw = max(len(label), len(key)) + 4
    if ftype == TA:
        raw = max(raw, 32)
    return max(15.0, min(40.0, float(raw)))


# ── vocabulary sheet ──────────────────────────────────────────────────────

def build_vocab_sheet(wb: Workbook) -> dict[str, str]:
    ws = wb.create_sheet("_vocabularies")
    ws.sheet_state = "hidden"
    ws.sheet_properties.tabColor = "888888"

    hdr_font  = Font(bold=True, color="444444", size=10)
    hdr_fill  = fill("DDDDDD")
    hdr_align = Alignment(horizontal="center", wrap_text=True)

    named_range_map: dict[str, str] = {}
    col = 1
    for vkey, values in VOCAB.items():
        col_letter = get_column_letter(col)
        hdr = ws.cell(row=1, column=col, value=vkey)
        hdr.font      = hdr_font
        hdr.fill      = hdr_fill
        hdr.alignment = hdr_align
        for r, val in enumerate(values, start=2):
            ws.cell(row=r, column=col, value=val)
        end_row  = len(values) + 1
        ref      = f"'_vocabularies'!${col_letter}$2:${col_letter}${end_row}"
        rng_name = f"vocab_{vkey}"
        named_range_map[vkey] = rng_name
        wb.defined_names[rng_name] = DefinedName(name=rng_name, attr_text=ref)
        ws.column_dimensions[col_letter].width = 32
        col += 1

    return named_range_map


# ── instructions sheet ────────────────────────────────────────────────────
# 7-column layout:
#  A = Sheet / bullet label   (~20)
#  B = Field Label             (~26)
#  C = Field Key               (~26)
#  D = Type                    (~22)
#  E = Definition              (~52)
#  F = Example Value           (~30)
#  G = Fill Instruction / Notes (~26)

_NCOLS_INSTR = 7   # number of columns used by Instructions sheet

def build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_properties.tabColor = HDR_BG

    # column widths
    col_widths = [20, 26, 26, 22, 52, 30, 26]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    last_col = get_column_letter(_NCOLS_INSTR)

    # ── local helpers ─────────────────────────────────────────────────────
    def merge_all(r: int) -> None:
        ws.merge_cells(f"A{r}:{last_col}{r}")

    def merge_text(r: int) -> None:
        """Merge B through last column for instruction body rows."""
        ws.merge_cells(f"B{r}:{last_col}{r}")

    def heading(r: int, text: str) -> None:
        merge_all(r)
        c = ws.cell(row=r, column=1, value=text)
        c.font      = Font(bold=True, size=13, color=HDR_BG)
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22

    def body(r: int, bullet: str, text: str) -> None:
        merge_text(r)
        ws.cell(row=r, column=1, value=bullet).alignment = \
            Alignment(horizontal="right", vertical="top")
        ws.cell(row=r, column=2, value=text).alignment = \
            Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 32

    def blank(r: int) -> None:
        ws.row_dimensions[r].height = 8

    # ── title ─────────────────────────────────────────────────────────────
    merge_all(1)
    t = ws.cell(row=1, column=1,
                value="ForestDSS Characterization Form \u2013 Excel Version")
    t.font      = Font(bold=True, size=16, color=HDR_FG)
    t.fill      = fill(HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    merge_all(2)
    sub = ws.cell(row=2, column=1,
                  value="COST Action CA22141 \u00b7 Training School Exercise")
    sub.font      = Font(italic=True, size=11, color="AAAAAA")
    sub.fill      = fill(HDR_BG)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    row = 3

    # ── how-to section ────────────────────────────────────────────────────
    blank(row); row += 1
    heading(row, "How to use this workbook"); row += 1
    for bullet, text in [
        ("1.", "Enter a DSS_ID in column A of every sheet. Use the DSS acronym or a short slug, e.g. SILVA-DSS. This ID links all instances across sections."),
        ("2.", "Add one row per instance. Multiple software components, models, or case studies each get their own row on the respective sheet."),
        ("3.", "Use the dropdown menus where available (cells with a small arrow). Selecting an invalid value will trigger an error message."),
        ("4.", "For multicheck fields, enter values separated by semicolons, e.g.:  Boreal; Continental; Atlantic. Header cells have a red-triangle comment with guidance."),
        ("5.", "Hover over any column header cell to read the field definition, fill instruction, and a Heureka example value."),
        ("6.", "Row 2 (grey, hidden) contains machine-readable field keys. Do not delete it \u2014 it is used for programmatic export."),
        ("7.", "The JSON export from the HTML form (forestdss_form.html) is preferred for wiki import. Use this Excel file as an offline data-collection aid."),
    ]:
        body(row, bullet, text)
        row += 1

    # ── sheets index ──────────────────────────────────────────────────────
    blank(row); row += 1
    heading(row, "Sheets in this workbook"); row += 1

    for ci, txt in [(1, "Sheet"), (2, "Contents")]:
        c = ws.cell(row=row, column=ci, value=txt)
        c.font = Font(bold=True, size=10)
        c.fill = fill(KEY_BG)
        c.alignment = Alignment(vertical="center")
    ws.merge_cells(f"B{row}:{last_col}{row}")
    ws.row_dimensions[row].height = 16
    row += 1

    for sname in SHEET_ORDER:
        ws.cell(row=row, column=1, value=sname).alignment = \
            Alignment(horizontal="left", vertical="top")
        ws.merge_cells(f"B{row}:{last_col}{row}")
        c = ws.cell(row=row, column=2, value=SHEET_DESCRIPTIONS[sname])
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 15
        row += 1

    blank(row)
    merge_all(row)
    note = ws.cell(row=row, column=1,
                   value="\u2139\ufe0f  The _vocabularies sheet (hidden) contains all "
                         "dropdown lists. Do not rename or delete it.")
    note.font      = Font(italic=True, color="666666", size=10)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 20
    row += 1

    # ── field reference table ─────────────────────────────────────────────
    blank(row); row += 1
    heading(row, "Field Reference \u2013 all sheets"); row += 1

    merge_all(row)
    intro = ws.cell(row=row, column=1,
                    value="One row per field across all 16 data sheets. "
                          "The same information appears as hover comments on "
                          "each column header cell in the data sheets. "
                          "Example values are drawn from the Heureka DSS (Sweden/SLU).")
    intro.font      = Font(italic=True, size=10, color="555555")
    intro.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 22
    row += 1

    # table header
    table_hdr_row = row
    col_headers = ["Sheet", "Field Label", "Field Key", "Type",
                   "Definition", "Example Value (Heureka)", "Fill Instruction"]
    for ci, txt in enumerate(col_headers, start=1):
        c = ws.cell(row=row, column=ci, value=txt)
        c.font      = Font(bold=True, color=HDR_FG, size=10)
        c.fill      = fill(HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[row].height = 22
    row += 1

    # data rows
    alt_fill_obj = fill(ALT_BG)
    wht_fill_obj = fill(WHT_BG)
    left_top     = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ri           = 0  # row counter for alternating colour

    for sname in SHEET_ORDER:
        for key, label, ftype in SHEET_FIELDS[sname]:
            meta = FIELD_META.get(key)
            ex_override = FIELD_EXAMPLES.get((sname, key))
            if meta:
                defn, default_ex = meta
                ex = ex_override if ex_override is not None else default_ex
            else:
                defn = ""
                ex = ex_override or ""

            instr    = _fill_instruction(ftype, key)
            type_lbl = _type_label(ftype)
            row_fill = alt_fill_obj if ri % 2 == 0 else wht_fill_obj

            values = [sname, label, key, type_lbl, defn, ex, instr]
            for ci, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=ci, value=val)
                c.fill      = row_fill
                c.alignment = left_top
                c.font      = Font(size=10)

            ws.row_dimensions[row].height = 28
            row += 1
            ri  += 1

    # auto-filter on the field table header row (only on Instructions sheet)
    ws.auto_filter.ref = \
        f"A{table_hdr_row}:{get_column_letter(_NCOLS_INSTR)}{table_hdr_row}"

    ws.freeze_panes = "A3"


# ── data sheet builder ────────────────────────────────────────────────────

def build_data_sheet(
    wb: Workbook,
    sheet_name: str,
    fields: list[tuple[str, str, str]],
    named_range_map: dict[str, str],
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = HDR_BG

    hdr_fill  = fill(HDR_BG)
    hdr_font  = Font(bold=True, color=HDR_FG, size=11)
    key_fill  = fill(KEY_BG)
    key_font  = Font(color=KEY_FG, size=10)
    alt_fill  = fill(ALT_BG)
    wht_fill  = fill(WHT_BG)
    ctr_wrap  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    ncols = len(fields)

    # ── row 1: human-readable headers + rich comments ─────────────────────
    ws.row_dimensions[1].height = 30
    for ci, (key, label, ftype) in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = ctr_wrap

        # build comment text from FIELD_META
        meta        = FIELD_META.get(key)
        ex_override = FIELD_EXAMPLES.get((sheet_name, key))
        if meta:
            defn, default_ex = meta
            ex = ex_override if ex_override is not None else default_ex
        else:
            defn = ""
            ex   = ex_override or ""

        instr = _fill_instruction(ftype, key)

        parts = []
        if defn:
            parts.append(f"Definition:\n{defn}")
        parts.append(f"How to fill:\n{instr}")
        if ex:
            parts.append(f"Example (Heureka):\n{ex}")
        cmt_text = "\n\n".join(parts)

        cmt         = Comment(cmt_text, "ForestDSS")
        cmt.width   = 310
        cmt.height  = 130 if defn else 80
        cell.comment = cmt

    # ── row 2: machine-readable keys (hidden) ────────────────────────────
    ws.row_dimensions[2].hidden = True
    ws.row_dimensions[2].height = 0
    for ci, (key, label, ftype) in enumerate(fields, start=1):
        cell = ws.cell(row=2, column=ci, value=key)
        cell.font      = key_font
        cell.fill      = key_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── rows 3 … 2+DATA_ROWS: empty data rows ────────────────────────────
    for ri in range(3, 3 + DATA_ROWS):
        row_fill = alt_fill if (ri % 2 == 1) else wht_fill
        ws.row_dimensions[ri].height = 18
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.fill      = row_fill
            cell.alignment = left_top

    # ── column widths ─────────────────────────────────────────────────────
    for ci, (key, label, ftype) in enumerate(fields, start=1):
        w = _col_width(label, key, ftype)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── data validation ───────────────────────────────────────────────────
    end_row = 2 + DATA_ROWS
    for ci, (key, label, ftype) in enumerate(fields, start=1):
        col_letter = get_column_letter(ci)
        cell_range = f"{col_letter}3:{col_letter}{end_row}"
        formula1   = None

        if ftype == B3:
            formula1 = "vocab_yesno"
        elif ftype.startswith("sel:"):
            vkey     = ftype[4:]
            formula1 = named_range_map.get(vkey, f"vocab_{vkey}")

        if formula1:
            dv = DataValidation(
                type="list",
                formula1=formula1,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid value",
                error="Please select a value from the dropdown list.",
            )
            ws.add_data_validation(dv)
            dv.add(cell_range)

    # ── freeze top row (no auto-filter on data sheets) ────────────────────
    ws.freeze_panes = "A3"


# ── main ──────────────────────────────────────────────────────────────────

def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Vocabulary sheet + named ranges
    named_range_map = build_vocab_sheet(wb)

    # 2. Instructions (at front)
    build_instructions_sheet(wb)

    # 3. Data sheets
    for sheet_name in SHEET_ORDER:
        build_data_sheet(
            wb, sheet_name, SHEET_FIELDS[sheet_name], named_range_map
        )

    # 4. Push _vocabularies to the very end
    names   = wb.sheetnames
    voc_pos = names.index("_vocabularies")
    wb.move_sheet("_vocabularies", offset=len(names) - 1 - voc_pos)

    # 5. Save
    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "forestdss_form.xlsx"
    )
    wb.save(out_path)
    print("Generated forestdss_form.xlsx")


if __name__ == "__main__":
    main()
